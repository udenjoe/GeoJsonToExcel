from collections import defaultdict
import json
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet
from openpyxl.chart import LineChart,Reference

def main(argv):
    fileName = sys.argv[1]
    print("File:", fileName)
    outFile = fileName.replace('.geojson', '.xlsx')
    wb = Workbook(write_only=False)
    chartSheet = wb.create_sheet("Charts", 0)

    with open(fileName, "r") as inputFile:
        json_content = json.load(inputFile)
        json_contents = str(json_content)
        json_contents = json_contents.replace("u'", "'")
        features = json_content['features']
        listValues = defaultdict(list)
        for feature in features:
            properties = feature['properties']
            property = properties.items()
            for key, value in property:
                print("key %s value %s" % (key, value))
                listValues[key].append(value)

        chartCount = 2


        for key, value in listValues.items():
            print(value)
            worksheet = wb.create_sheet(key)
            worksheet.title = key
            worksheet = wb[key]
            #first row is titles
            worksheet.cell(row=1, column=1, value='timestamp')
            worksheet.cell(row=1, column=2, value='value')
            rowcount = 2 #1 indexed start at second row
            for item in value:
                worksheet.cell(row=rowcount, column=1, value=rowcount - 2)#timestamp start at 0
                insertValue = item
                if type(item) is bool:
                    if(item):
                        insertValue = 1
                    else:
                        insertValue = 0
                worksheet.cell(row=rowcount, column=2, value=insertValue)
                rowcount += 1
            values = Reference(worksheet, min_col = 2, min_row = 2, max_col = 2, max_row = rowcount)
            chart = LineChart()
            chart.add_data(values)
            chart.title = key
            chart.height = 10
            chart.width = 50
            chart.x_axis.title = "timestamp"
            chart.y_axis.title = "values"
            chart.legend = None
            location = 'B%d' % chartCount
            chartSheet.add_chart(chart, location)
            chartCount += 19
        
        sheet = wb["Sheet"]
        wb.remove(sheet)
        wb.save(outFile)     


if __name__ == "__main__":
   main(sys.argv[1:])
